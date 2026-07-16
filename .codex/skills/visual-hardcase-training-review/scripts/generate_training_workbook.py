#!/usr/bin/env python3
"""Generate a local newcomer training workbook for visual hardcase work."""

from __future__ import annotations

import argparse
import json
import os
from datetime import datetime
from pathlib import Path
from typing import Any
import urllib.error
import urllib.request
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


DEFAULT_PUBLIC_WIKI_URL = "https://wiki.muchenai.com"
DEFAULT_BASE_URL = os.environ.get("LLM_WIKI_API_BASE_URL", DEFAULT_PUBLIC_WIKI_URL).rstrip("/")
DEFAULT_PROJECT_ID = os.environ.get(
    "LLM_WIKI_PROJECT_ID",
    os.environ.get("LLM_WIKI_PROJECT_PATH", "7ad8995a9c34304f"),
)
DEFAULT_PROJECT_PATH = DEFAULT_PROJECT_ID

TASK_TYPES = [
    ("基础感知", "实物判断", "主图 + 至少7个选项", "同款/同物判断；图片复杂且选项有干扰；答案输出编号。"),
    ("基础感知", "多图变化", "多图，通常>=7个变化选项", "找出可见变化；选项描述必须具体；答案按要求输出字母。"),
    ("基础感知", "读示数", "单图或少量图", "仪表刻度清晰；精度、单位和读数口径唯一。"),
    ("时空间理解", "室内方位推理", "多图", "必须给方向基准；问题依赖室内空间关系。"),
    ("时空间理解", "多图动态排序", "多图", "排序逻辑唯一；推理要写视觉线索。"),
    ("时空间理解", "多图去重计数", "多图", "同一空间多角度去重；反光/投影不计入。"),
    ("推理游戏", "齿轮", "单图或多图", "避免只问一个齿轮；通常至少问两个齿轮方向/状态。"),
    ("推理游戏", "漫画排序", "至少5张", "情节逻辑唯一；推理不能只复述答案。"),
    ("推理游戏", "地图规划", "地图/街景组合", "必须有方向基准或在prompt中说明假设。"),
    ("推理游戏", "图形关系", "多图形，选项>=7", "闭合图形、关系唯一；避免外框干扰和选项歧义。"),
    ("综合推理感知", "多图场景判别", "主图 + 至少7个选项", "找同场景且满足指定物体/条件；选项有干扰。"),
    ("综合推理感知", "双目双图结合", "双图", "两张局部图需要结合；答案可有特例但必须说明口径。"),
    ("综合推理感知", "多图同屋判别", "多图", "判断同房间；重叠不足时prompt需补充口径。"),
    ("综合推理感知", "判断与反思-纠错", "4组图文问答", "四个子问题类型不同；定位唯一错误项。"),
    ("OCR", "OCR票据", "票据/小票/发票", "真实票据；隐私脱敏；只提取文字/金额/日期，不做算术。"),
]

ENTRY_COLUMNS = [
    "case_id",
    "新人姓名",
    "训练批次",
    "题型分类",
    "图片数量",
    "图片文件夹或图片链接",
    "图片编号说明",
    "prompt",
    "answer",
    "答案格式限定是否已写入prompt",
    "答案是否唯一",
    "推理过程",
    "信源链接或标注图路径",
    "是否需要验证结果",
    "answer(验证结果)",
    "模型是否答错",
    "初标是否完成",
    "自检备注",
]

SELF_CHECK_ITEMS = [
    "题型属于当前v3有效题型，不是历史停用题型。",
    "图片清晰、可检查，关键细节没有被遮挡或过度压缩。",
    "多图任务已在图片内部或说明中标清编号，prompt不引用文件名。",
    "prompt明确依赖图片信息，不能只靠常识回答。",
    "prompt已严格约束answer输出格式。",
    "answer客观唯一；如存在等价表达，已用“或”列出。",
    "需要推理的复杂题型已写出视觉线索，不是只复述答案。",
    "隐私、人脸、手机号、证件号、广告水印等风险已处理。",
    "方向/路线/空间题已写明方向基准或必要假设。",
    "模型验证结果显示目标模型答错，且验证结果字段已保留。",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate visual hardcase newcomer training workbook.")
    parser.add_argument("--output-dir", help="Output directory. Defaults to ~/Downloads/visual-hardcase-training/<timestamp>.")
    parser.add_argument("--filename", default="training-workbook.xlsx", help="Workbook filename.")
    parser.add_argument("--base-url", default=DEFAULT_BASE_URL, help="LLM Wiki API base URL.")
    parser.add_argument(
        "--project-id",
        "--project-path",
        dest="project_id",
        default=DEFAULT_PROJECT_ID,
        help="LLM Wiki project id. --project-path is kept as a compatibility alias.",
    )
    parser.add_argument("--timeout", type=float, default=8.0, help="HTTP timeout seconds.")
    return parser.parse_args()


def build_wiki_headers() -> dict[str, str]:
    headers = {"Content-Type": "application/json"}
    token = os.environ.get("LLM_WIKI_API_TOKEN", "").strip()
    if token:
        headers["Authorization"] = f"Bearer {token}"
    return headers


def search_wiki(base_url: str, project_id: str, query: str, timeout: float) -> list[dict[str, Any]]:
    payload = {
        "query": query,
        "topK": 5,
        "includeContent": False,
    }
    body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    encoded_project_id = quote(project_id.strip(), safe="")
    req = urllib.request.Request(
        f"{base_url.rstrip('/')}/api/v1/projects/{encoded_project_id}/search",
        data=body,
        headers=build_wiki_headers(),
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        data = json.loads(resp.read().decode("utf-8"))
    return list(data.get("results") or [])


def collect_wiki_evidence(args: argparse.Namespace) -> tuple[str, list[dict[str, str]]]:
    queries = [
        "v3题型总览 准入标准",
        "模型验证 自检流程",
        "难题构造 通用 避坑",
        "作业步骤 构造思路 prompt answer",
    ]
    evidence: list[dict[str, str]] = []
    try:
        for query in queries:
            for item in search_wiki(args.base_url, args.project_id, query, args.timeout):
                title = str(item.get("title") or item.get("path") or "")
                path = str(item.get("path") or "")
                if path and not any(existing["path"] == path for existing in evidence):
                    evidence.append({"title": title, "path": path})
        return "LLM Wiki API 已连接", evidence[:12]
    except (urllib.error.URLError, urllib.error.HTTPError, TimeoutError, json.JSONDecodeError) as exc:
        return f"LLM Wiki API 未连接，当前表格为内置兜底模板: {exc}", []


def set_title(row) -> None:
    for cell in row:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="305496")
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def add_instructions(ws, wiki_status: str, evidence: list[dict[str, str]]) -> None:
    rows = [
        ["模块", "说明"],
        ["项目目标", "构造高质量、高难度、答案唯一的视觉理解hardcase，用于检验模型在图像+prompt下的能力边界。"],
        ["核心流程", "确认v3题型 -> 采集图片/素材 -> 写prompt -> 写answer -> 跑模型验证 -> 自检 -> 提交/质检。"],
        ["构造公式", "基于分类找图片，基于图片写prompt，结合图片+prompt设置answer。"],
        ["模型验证", "是否需要验证结果通常填“是”；目标模型答错才说明hardcase难度合格。"],
        ["新人操作", "只填写“训练题填写”sheet；每一行是一道训练题。填完后交给review脚本生成“质检结果”。"],
        ["API状态", wiki_status],
    ]
    for row in rows:
        ws.append(row)
    ws.append([])
    ws.append(["知识库依据", "路径"])
    for item in evidence:
        ws.append([item["title"], item["path"]])
    set_title(ws[1])
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 92


def add_entry_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("训练题填写")
    ws.append(ENTRY_COLUMNS)
    set_title(ws[1])
    widths = {
        "A": 16, "B": 14, "C": 14, "D": 20, "E": 14, "F": 36,
        "G": 24, "H": 60, "I": 36, "J": 18, "K": 14, "L": 50,
        "M": 32, "N": 18, "O": 36, "P": 14, "Q": 14, "R": 36,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    for row_num in range(2, 52):
        ws.cell(row=row_num, column=1).value = f"case-{row_num - 1:03d}"
    for row in ws.iter_rows(min_row=1, max_row=52, min_col=1, max_col=len(ENTRY_COLUMNS)):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws["H1"].comment = Comment("prompt必须包含输出格式要求，例如：请只输出选项字母，按从左到右顺序，不加分隔符。", "Codex")
    ws["F1"].comment = Comment("本地表格可填写文件夹路径、云文档链接，或图片URL列表。URL列表建议一行一张图。", "Codex")
    ws["O1"].comment = Comment("保留模型原始验证结果。若模型答对，应修改题目或换图。", "Codex")


def add_task_reference_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("题型规则速查")
    ws.append(["大类", "题型", "图片/素材要求", "关键约束"])
    set_title(ws[1])
    for row in TASK_TYPES:
        ws.append(list(row))
    for col, width in {"A": 18, "B": 22, "C": 30, "D": 72}.items():
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def add_checklist_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("自检清单")
    ws.append(["序号", "检查项", "是否通过", "备注"])
    set_title(ws[1])
    for index, item in enumerate(SELF_CHECK_ITEMS, start=1):
        ws.append([index, item, "", ""])
    for col, width in {"A": 8, "B": 86, "C": 14, "D": 40}.items():
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def add_result_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("质检结果")
    ws.append(["case_id", "结论", "总分", "阻断项", "主要问题", "修改建议", "规则依据", "评分时间"])
    set_title(ws[1])
    for col, width in {"A": 16, "B": 14, "C": 10, "D": 42, "E": 52, "F": 52, "G": 40, "H": 22}.items():
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def add_options_and_validations(wb: Workbook) -> None:
    options = wb.create_sheet("_options")
    options.sheet_state = "hidden"
    options.append(["task_types", "yes_no", "validation_required", "model_wrong"])
    for index, (_, task_type, *_rest) in enumerate(TASK_TYPES, start=2):
        options.cell(index, 1).value = task_type
    for index, value in enumerate(["是", "否"], start=2):
        options.cell(index, 2).value = value
    for index, value in enumerate(["是", "否"], start=2):
        options.cell(index, 3).value = value
    for index, value in enumerate(["是", "否", "待验证"], start=2):
        options.cell(index, 4).value = value

    entry = wb["训练题填写"]
    task_dv = DataValidation(type="list", formula1=f"=_options!$A$2:$A${len(TASK_TYPES) + 1}")
    yes_no_dv = DataValidation(type="list", formula1="=_options!$B$2:$B$3")
    checklist_yes_no_dv = DataValidation(type="list", formula1="=_options!$B$2:$B$3")
    model_dv = DataValidation(type="list", formula1="=_options!$D$2:$D$4")
    for dv in (task_dv, yes_no_dv, model_dv):
        entry.add_data_validation(dv)
    task_dv.add("D2:D51")
    yes_no_dv.add("J2:K51")
    yes_no_dv.add("N2:N51")
    yes_no_dv.add("Q2:Q51")
    model_dv.add("P2:P51")

    checklist = wb["自检清单"]
    checklist.add_data_validation(checklist_yes_no_dv)
    checklist_yes_no_dv.add("C2:C50")


def main() -> int:
    args = parse_args()
    if args.output_dir:
        output_dir = Path(args.output_dir).expanduser()
    else:
        stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        output_dir = Path.home() / "Downloads" / "visual-hardcase-training" / stamp
    output_dir.mkdir(parents=True, exist_ok=True)
    workbook_path = output_dir / args.filename

    wiki_status, evidence = collect_wiki_evidence(args)

    wb = Workbook()
    ws = wb.active
    ws.title = "填写说明"
    add_instructions(ws, wiki_status, evidence)
    add_entry_sheet(wb)
    add_task_reference_sheet(wb)
    add_checklist_sheet(wb)
    add_result_sheet(wb)
    add_options_and_validations(wb)
    wb.save(workbook_path)

    print(f"created: {workbook_path}")
    print(wiki_status)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
