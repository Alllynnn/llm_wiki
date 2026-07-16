#!/usr/bin/env python3
"""Review a filled visual hardcase newcomer training workbook."""

from __future__ import annotations

import argparse
import json
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Any
import urllib.error
import urllib.request
from urllib.parse import quote

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


DEFAULT_PUBLIC_WIKI_URL = "https://wiki.muchenai.com"
DEFAULT_BASE_URL = os.environ.get("LLM_WIKI_API_BASE_URL", DEFAULT_PUBLIC_WIKI_URL).rstrip("/")
DEFAULT_PROJECT_ID = os.environ.get(
    "LLM_WIKI_PROJECT_ID",
    os.environ.get("LLM_WIKI_PROJECT_PATH", "7ad8995a9c34304f"),
)
DEFAULT_PROJECT_PATH = DEFAULT_PROJECT_ID

ACTIVE_TASK_TYPES = {
    "实物判断",
    "多图变化",
    "读示数",
    "室内方位推理",
    "多图动态排序",
    "多图去重计数",
    "齿轮",
    "漫画排序",
    "地图规划",
    "图形关系",
    "多图场景判别",
    "双目双图结合",
    "多图同屋判别",
    "判断与反思-纠错",
    "OCR票据",
}

COMPLEX_REASONING_TYPES = {
    "实物判断",
    "多图变化",
    "室内方位推理",
    "多图动态排序",
    "多图去重计数",
    "漫画排序",
    "地图规划",
    "图形关系",
    "多图场景判别",
    "双目双图结合",
    "多图同屋判别",
    "判断与反思-纠错",
}

MULTI_IMAGE_TYPES = {
    "实物判断",
    "多图变化",
    "室内方位推理",
    "多图动态排序",
    "多图去重计数",
    "漫画排序",
    "多图场景判别",
    "双目双图结合",
    "多图同屋判别",
    "判断与反思-纠错",
}

UNREADABLE_IMAGE_MARKERS = (
    "云表格图片列未读取到文本",
    "请人工查看单元格图片",
)

REQUIRED_COLUMNS = [
    "case_id",
    "题型分类",
    "图片文件夹或图片链接",
    "prompt",
    "answer",
    "答案格式限定是否已写入prompt",
    "答案是否唯一",
    "是否需要验证结果",
    "answer(验证结果)",
    "模型是否答错",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Review a filled newcomer training workbook.")
    parser.add_argument("workbook", help="Path to training-workbook.xlsx")
    parser.add_argument("--output", help="Reviewed workbook path. Defaults to <input>-reviewed.xlsx.")
    parser.add_argument("--report", help="Markdown report path. Defaults to <input>-review-report.md.")
    parser.add_argument("--base-url", default=DEFAULT_BASE_URL, help="LLM Wiki API base URL.")
    parser.add_argument(
        "--project-id",
        "--project-path",
        dest="project_id",
        default=DEFAULT_PROJECT_ID,
        help="LLM Wiki project id. --project-path is kept as a compatibility alias.",
    )
    parser.add_argument("--timeout", type=float, default=8.0, help="HTTP timeout seconds.")
    return parser.parse_args()


def cell_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def build_wiki_headers() -> dict[str, str]:
    headers = {"Content-Type": "application/json"}
    token = os.environ.get("LLM_WIKI_API_TOKEN", "").strip()
    if token:
        headers["Authorization"] = f"Bearer {token}"
    return headers


def search_wiki(base_url: str, project_id: str, query: str, timeout: float) -> list[dict[str, Any]]:
    payload = {
        "query": query,
        "topK": 3,
        "includeContent": False,
    }
    body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    encoded_project_id = quote(project_id.strip(), safe="")
    req = urllib.request.Request(
        f"{base_url.rstrip('/')}/api/v1/projects/{encoded_project_id}/search",
        data=body,
        headers=build_wiki_headers(),
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        data = json.loads(resp.read().decode("utf-8"))
    return list(data.get("results") or [])


def has_format_constraint(prompt: str) -> bool:
    keywords = [
        "仅输出",
        "只输出",
        "输出格式",
        "请按",
        "大写输出",
        "小写输出",
        "不要输出",
        "不加",
        "无需符号",
        "无需空格",
        "无需换行",
        "不带",
        "只用",
        "格式为",
        "答案格式",
        "示例回复",
        "阿拉伯数字",
        "顿号隔开",
        "逗号分隔",
    ]
    patterns = [
        r"(?:按照|按).{0,60}输出",
        r"输出.{0,30}(?:无需|不加|不带|不要)",
    ]
    return any(keyword in prompt for keyword in keywords) or any(re.search(pattern, prompt) for pattern in patterns)


def is_binary_prompt(prompt: str) -> bool:
    patterns = [r"是否", r"对吗", r"正确吗", r"是还是否", r"回答是或否", r"判断.*对错"]
    return any(re.search(pattern, prompt) for pattern in patterns)


def cap_score(score: int, cap: int) -> int:
    return min(score, cap)


def review_row(row: dict[str, str], args: argparse.Namespace) -> dict[str, Any]:
    score = 100
    blockers: list[str] = []
    issues: list[str] = []
    suggestions: list[str] = []
    evidence: list[str] = []
    api_unavailable = False

    require_model_validation = getattr(args, "require_model_validation", True)
    required_columns = list(REQUIRED_COLUMNS)
    if not require_model_validation:
        required_columns = [
            column for column in required_columns
            if column not in {"是否需要验证结果", "answer(验证结果)", "模型是否答错"}
        ]
    for column in required_columns:
        if not row.get(column):
            blockers.append(f"缺少必填字段：{column}")
            score = cap_score(score, 50)

    task_type = row.get("题型分类", "")
    prompt = row.get("prompt", "")
    answer = row.get("answer", "")
    image_ref = row.get("图片文件夹或图片链接", "")
    image_count = row.get("图片数量", "")
    image_labels = row.get("图片编号说明", "")
    reasoning = row.get("推理过程", "")

    if task_type and task_type not in ACTIVE_TASK_TYPES:
        blockers.append(f"题型不是当前v3有效题型：{task_type}")
        score = cap_score(score, 45)

    if not image_ref or image_count == "0" or any(marker in image_ref for marker in UNREADABLE_IMAGE_MARKERS):
        blockers.append("图片材料缺失或未被脚本读取，不能确认视觉证据。")
        suggestions.append("提供可访问的图片链接/文件夹，或先把云表单元格图片导出为可读取 URL。")
        score = cap_score(score, 50)
    elif image_ref and re.search(r"\.(png|jpg|jpeg|webp|gif|bmp)$", image_ref, re.I):
        issues.append("图片字段看起来是单个文件；若是多图题，需要提供文件夹或完整图片列表。")
        score -= 5

    if task_type in MULTI_IMAGE_TYPES and not image_labels:
        blockers.append("多图题缺少图片编号说明，prompt/answer容易引用不稳定文件名。")
        score = cap_score(score, 65)

    format_flag = row.get("答案格式限定是否已写入prompt", "")
    if format_flag != "是" or not has_format_constraint(prompt):
        blockers.append("prompt未明确约束answer输出格式。")
        suggestions.append("在prompt末尾增加“请只输出...，不要输出解释/单位/分隔符”等格式约束。")
        score = cap_score(score, 70)

    unique_flag = row.get("答案是否唯一", "")
    if unique_flag != "是" and task_type != "双目双图结合":
        blockers.append("答案唯一性未确认。")
        score = cap_score(score, 60)

    need_validation = row.get("是否需要验证结果", "")
    model_wrong = row.get("模型是否答错", "")
    validation_text = row.get("answer(验证结果)", "")
    if require_model_validation:
        if need_validation != "是":
            blockers.append("hardcase训练题通常必须进行模型验证。")
            score = cap_score(score, 65)
        if model_wrong != "是" or not validation_text:
            blockers.append("模型验证未证明目标模型答错。")
            score = cap_score(score, 60)
    elif not validation_text:
        issues.append("当前云表模式未强制目标模型验证；正式hardcase提交前仍需补模型验证或接入公网QC API。")
        score -= 3

    if task_type in COMPLEX_REASONING_TYPES and len(reasoning) < 20:
        issues.append("复杂题型的推理过程过短，可能被判为敷衍或不可复核。")
        suggestions.append("补充可观察视觉线索，以及这些线索如何推出answer。")
        score -= 12

    if is_binary_prompt(prompt) and task_type not in {"判断与反思-纠错"}:
        issues.append("prompt疑似二元判断题，通常不符合v3难题构造要求。")
        score -= 10

    if len(answer) >= 2 and len(prompt) > 0 and answer in prompt:
        issues.append("answer疑似直接出现在prompt中，需要人工确认是否泄题。")
        score -= 8

    if not row.get("信源链接或标注图路径", ""):
        issues.append("缺少信源链接或标注图路径，质检可追溯性不足。")
        score -= 6

    try:
        if task_type:
            results = search_wiki(args.base_url, args.project_id, f"{task_type} 规则 bad case 质检", args.timeout)
        else:
            results = search_wiki(args.base_url, args.project_id, "v3题型 准入 质检", args.timeout)
        for item in results:
            title = cell_text(item.get("title") or item.get("path"))
            path = cell_text(item.get("path"))
            if title or path:
                evidence.append(f"{title} ({path})")
    except (urllib.error.URLError, urllib.error.HTTPError, TimeoutError, json.JSONDecodeError) as exc:
        api_unavailable = True
        blockers.append(f"无法连接LLM Wiki API，不能给出正式评分：{exc}")
        score = cap_score(score, 50)

    if api_unavailable:
        conclusion = "无法评分"
    elif blockers:
        conclusion = "不合格" if score < 70 else "需修改"
    elif issues:
        conclusion = "需修改" if score < 85 else "基本合格"
    else:
        conclusion = "合格"

    score = max(0, min(100, score))
    return {
        "case_id": row.get("case_id") or "",
        "conclusion": conclusion,
        "score": score,
        "blockers": blockers,
        "issues": issues,
        "suggestions": suggestions,
        "evidence": evidence,
    }


def read_submissions(workbook_path: Path) -> tuple[Any, list[dict[str, str]]]:
    wb = load_workbook(workbook_path)
    if "训练题填写" not in wb.sheetnames:
        raise SystemExit("Workbook is missing sheet: 训练题填写")
    ws = wb["训练题填写"]
    headers = [cell_text(cell.value) for cell in ws[1]]
    missing = [column for column in REQUIRED_COLUMNS if column not in headers]
    if missing:
        raise SystemExit(f"Workbook is missing required columns: {', '.join(missing)}")

    submissions: list[dict[str, str]] = []
    content_headers = [header for header in headers if header != "case_id"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        values = {headers[index]: cell_text(value) for index, value in enumerate(row) if index < len(headers)}
        if any(values.get(column, "") for column in content_headers):
            submissions.append(values)
    return wb, submissions


def ensure_result_sheet(wb: Any) -> Any:
    if "质检结果" in wb.sheetnames:
        ws = wb["质检结果"]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet("质检结果")
    headers = ["case_id", "结论", "总分", "阻断项", "主要问题", "修改建议", "规则依据", "评分时间"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="305496")
    for col, width in {"A": 16, "B": 14, "C": 10, "D": 42, "E": 52, "F": 52, "G": 50, "H": 22}.items():
        ws.column_dimensions[col].width = width
    return ws


def write_results(ws: Any, results: list[dict[str, Any]]) -> None:
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for result in results:
        ws.append([
            result["case_id"],
            result["conclusion"],
            result["score"],
            "\n".join(result["blockers"]) or "无明显阻断项",
            "\n".join(result["issues"]) or "无明显主要问题",
            "\n".join(result["suggestions"]) or "无",
            "\n".join(result["evidence"]) or "未检索到依据",
            now,
        ])
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def write_report(report_path: Path, workbook_path: Path, results: list[dict[str, Any]]) -> None:
    total = len(results)
    counts: dict[str, int] = {}
    for result in results:
        counts[result["conclusion"]] = counts.get(result["conclusion"], 0) + 1
    lines = [
        "# 新人训练评审报告",
        "",
        f"- Source workbook: `{workbook_path}`",
        f"- Total rows: {total}",
        f"- Generated at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        "",
        "## 批次概览",
        "",
    ]
    for key in ["合格", "基本合格", "需修改", "不合格", "无法评分"]:
        lines.append(f"- {key}: {counts.get(key, 0)}")
    lines.extend(["", "## 明细", ""])
    for result in results:
        lines.extend([
            f"### {result['case_id'] or '未命名case'}",
            "",
            f"- 结论: {result['conclusion']}",
            f"- 总分: {result['score']}",
            f"- 阻断项: {'；'.join(result['blockers']) if result['blockers'] else '无明显阻断项'}",
            f"- 主要问题: {'；'.join(result['issues']) if result['issues'] else '无明显主要问题'}",
            f"- 修改建议: {'；'.join(result['suggestions']) if result['suggestions'] else '无'}",
            f"- 规则依据: {'；'.join(result['evidence']) if result['evidence'] else '未检索到依据'}",
            "",
        ])
    report_path.write_text("\n".join(lines), encoding="utf-8")


def main() -> int:
    args = parse_args()
    workbook_path = Path(args.workbook).expanduser().resolve()
    output_path = Path(args.output).expanduser().resolve() if args.output else workbook_path.with_name(f"{workbook_path.stem}-reviewed.xlsx")
    report_path = Path(args.report).expanduser().resolve() if args.report else workbook_path.with_name(f"{workbook_path.stem}-review-report.md")

    wb, submissions = read_submissions(workbook_path)
    if not submissions:
        raise SystemExit("No filled rows found in 训练题填写.")
    results = [review_row(row, args) for row in submissions]
    result_ws = ensure_result_sheet(wb)
    write_results(result_ws, results)
    wb.save(output_path)
    write_report(report_path, workbook_path, results)

    print(f"reviewed_workbook: {output_path}")
    print(f"report: {report_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
