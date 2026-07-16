#!/usr/bin/env python3
"""Export a local JSON payload preview for a future public QC API."""

from __future__ import annotations

import argparse
import json
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export training workbook rows to a disabled API payload preview.")
    parser.add_argument("workbook", help="Path to reviewed or filled training workbook.")
    parser.add_argument("--out", help="Output JSON path. Defaults to <workbook>-payload-preview.json.")
    return parser.parse_args()


def cell_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def read_sheet(wb: Any, sheet_name: str) -> list[dict[str, str]]:
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    headers = [cell_text(cell.value) for cell in ws[1]]
    rows: list[dict[str, str]] = []
    content_headers = [header for header in headers if header != "case_id"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        item = {headers[index]: cell_text(value) for index, value in enumerate(row) if index < len(headers)}
        if any(item.get(header, "") for header in content_headers):
            rows.append(item)
    return rows


def main() -> int:
    args = parse_args()
    workbook_path = Path(args.workbook).expanduser().resolve()
    output_path = Path(args.out).expanduser().resolve() if args.out else workbook_path.with_name(f"{workbook_path.stem}-payload-preview.json")
    wb = load_workbook(workbook_path, data_only=True)
    payload = {
        "api_status": "disabled_until_contract_provided",
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "source_workbook": str(workbook_path),
        "submissions": read_sheet(wb, "训练题填写"),
        "review_results": read_sheet(wb, "质检结果"),
    }
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"payload_preview: {output_path}")
    print("upload_status: disabled_until_contract_provided")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
